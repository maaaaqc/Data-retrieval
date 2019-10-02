from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, fills
from openpyxl.styles.borders import Border, Side

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

title_background = PatternFill(patternType=fills.FILL_SOLID,
                               start_color='00CAE3FF')


def convert_excel(dictionary, excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "INDEX"
    c = ws.cell(1, 1, value="TITLE")
    c.border = thin_border
    c.font = Font(bold=True)
    c.fill = title_background
    c = ws.cell(1, 2, value="LINK")
    c.border = thin_border
    c.font = Font(bold=True)
    c.fill = title_background
    for key in dictionary.keys():
        wst = wb.create_sheet(key.upper())
        if not key == "nlpsinfo" \
            and not key == "profile" \
            and not "ccss" in key:
            c = wst.cell(1, 1, value=key.upper())
            c.border = thin_border
            c.font = Font(bold=True)
            c.fill = title_background
            index = 2
            for val in dictionary[key]:
                wst.cell(index, 1, value=val)
                index += 1
        else:
            index = 1
            for val in dictionary[key]:
                column = 1
                for i in val:
                    c = wst.cell(index, column, value=i)
                    if index == 1:
                        c.border = thin_border
                        c.font = Font(bold=True)
                        c.fill = title_background
                    column += 1
                index += 1
        index = 2
        for w in wb.worksheets:
            if not w.title == "INDEX":
                wb.worksheets[0][f"A{index}"].value = f"{w.title}"
                wb.worksheets[0][f"B{index}"].hyperlink = f"#'{w.title}'!A1"
                wb.worksheets[0][f"B{index}"].value = "Click here"
                wb.worksheets[0][f"B{index}"].style = "Hyperlink"
                index += 1
            for row in w.rows:
                for cell in row:
                    if cell.value:
                        if w.column_dimensions[cell.column_letter].width < len(str(cell.value))+8:
                            w.column_dimensions[cell.column_letter].width = len(str(cell.value))+8
    wb.save(str(excel))